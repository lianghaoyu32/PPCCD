import csv

import numpy as np
import openpyxl as ox
import pandas as pd
import math
import re
import os

'''读取文件名函数'''


def file_name(file_dir):
    for root, dirs, files in os.walk(file_dir):
        return files  # 当前文件夹下的所有文件的名称


'''获取文件夹名字的函数'''


def getfilename(filename):
    for root, dirs, files in os.walk(filename):
        array = dirs  # 当前文件夹下的所有子文件夹的名称
        if array:
            return array


# file_a = 'E:/研一/Community-Detection/dataset/20200605-M-2019153LC-The First Affiliated Hospital of Guangzhou Medical University - Liang Wenhua/Panel 1/P2017004-P117-2-T5-1-20200217/'
# file_a = 'E:/研一/Community-Detection/dataset/20200605-M-2019153LC-The First Affiliated Hospital of Guangzhou Medical University - Liang Wenhua/Panel 1/P2017004-P117-2-T5-2-20200217/'
# file_a = 'E:/研一/Community-Detection/dataset/20200605-M-2019153LC-The First Affiliated Hospital of Guangzhou Medical University - Liang Wenhua/Panel 2/117-3-T5-1/'
# file_a = 'E:/研一/Community-Detection/dataset/20200605-M-2019153LC-The First Affiliated Hospital of Guangzhou Medical University - Liang Wenhua/Panel 2/117-3-T5-2/'
file_a = 'E:/研一/Community-Detection/dataset/20200605-M-2019153LC-The First Affiliated Hospital of Guangzhou Medical University - Liang Wenhua/Panel 2/117-3-T5-3/'


line = 2

# A = 'CD4'
# B = 'CD20'
# C = 'CD38'
# D = 'CD66B'
# E = 'FOXP3'

A = 'CD8'
B = 'CD68'
C = 'CD133'
D = 'CD163'
E = 'PDL1'

wb = ox.Workbook()
ws = wb.active
ws['A1'] = '图片名称'
ws['B1'] = '社区数量'
ws['C1'] = A + '细胞数目'
ws['D1'] = B + '细胞数目'
ws['E1'] = C + '细胞数目'
ws['F1'] = D + '细胞数目'
ws['G1'] = E + '细胞数目'
ws['H1'] = A + '细胞百分比'
ws['I1'] = B + '细胞百分比'
ws['J1'] = C + '细胞百分比'
ws['K1'] = D + '细胞百分比'
ws['L1'] = E + '细胞百分比'
ws['M1'] = A + '-' + B + '连接数'
ws['N1'] = A + '-' + C + '连接数'
ws['O1'] = A + '-' + D + '连接数'
ws['P1'] = A + '-' + E + '连接数'
ws['Q1'] = B + '-' + C + '连接数'
ws['R1'] = B + '-' + D + '连接数'
ws['S1'] = B + '-' + E + '连接数'
ws['T1'] = C + '-' + D + '连接数'
ws['U1'] = C + '-' + E + '连接数'
ws['V1'] = D + '-' + E + '连接数'
ws['W1'] = A + '-' + A + '连接数'
ws['X1'] = B + '-' + B + '连接数'
ws['Y1'] = C + '-' + C + '连接数'
ws['Z1'] = D + '-' + D + '连接数'
ws['AA1'] = E + '-' + E + '连接数'
ws['AB1'] = A + '-' + B + '平均连接距离'
ws['AC1'] = A + '-' + C + '平均连接距离'
ws['AD1'] = A + '-' + D + '平均连接距离'
ws['AE1'] = A + '-' + E + '平均连接距离'
ws['AF1'] = B + '-' + C + '平均连接距离'
ws['AG1'] = B + '-' + D + '平均连接距离'
ws['AH1'] = B + '-' + E + '平均连接距离'
ws['AI1'] = C + '-' + D + '平均连接距离'
ws['AJ1'] = C + '-' + E + '平均连接距离'
ws['AK1'] = D + '-' + E + '平均连接距离'
ws['AL1'] = A + '-' + A + '平均连接距离'
ws['AM1'] = B + '-' + B + '平均连接距离'
ws['AN1'] = C + '-' + C + '平均连接距离'
ws['AO1'] = D + '-' + D + '平均连接距离'
ws['AP1'] = E + '-' + E + '平均连接距离'
ws['AQ1'] = '患者id'

'''获取文件夹名字,生成列表并进行遍历'''
big_file = getfilename(file_a)

'''文件夹个数(文件夹列表中的序号)'''
big_file_num = 0

for big_file_name in big_file:
    print("循环文件夹")
    print(big_file_name)
    '''去掉DATA字符，只留下前几位字符'''
    small_file = big_file_name[0:len(big_file_name) - 5]

    '''文件路径'''

    csv_l = file_name(file_a + big_file_name + '/' + big_file_name + '/')

    '''文件名匹配并读取'''
    CD4_file_name = []
    CD20_file_name = []
    CD38_file_name = []
    CD66B_file_name = []
    FOXP3_file_name = []
    csv_file_name = []  # 储存csv文件

    csv_pattern = re.compile('.+csv')  # 在文件夹中找出csv文件

    for csv_i in csv_l:
        csv_result = csv_pattern.findall(csv_i)  # 在字符串 csv_i 中查找与正则表达式模式 csv_pattern 匹配的所有子串
        if csv_result != []:
            csv_file_name.append(csv_result)

    for csv_name in csv_file_name:
        '''遍历csv文件名'''
        # df_csv = pd.read_csv(file_a + big_file_name + '/' + big_file_name + '/' + str(csv_name[0]), usecols=[1])

        Name = str(csv_pattern.findall(csv_name[0])[0]).rstrip('_DAPI_path_view.csv')
        print('Name:', Name)
        model_name = 'rber_pots' + '算法'  # 社区发现算法
        panel = 2  # 不同panel，不同的标志物
        # panel = 1  # 不同panel，不同的标志物

        '''查找病人ID'''
        q = Name + '_cell_seg_data.txt'
        # print(q)
        # chip_excel = pd.read_excel(
        #     "E:/研一/Community-Detection/dataset/20200605-M-2019153LC-The First Affiliated Hospital of Guangzhou Medical University - Liang Wenhua/(Panel_1)芯片住院信息对应.xlsx",
        #     usecols=[0])  # 芯片名称
        # patient_id = pd.read_excel(
        #     "E:/研一/Community-Detection/dataset/20200605-M-2019153LC-The First Affiliated Hospital of Guangzhou Medical University - Liang Wenhua/(Panel_1)芯片住院信息对应.xlsx",
        #     usecols=[1])  # 患者ID

        chip_excel = pd.read_excel(
            "E:/研一/Community-Detection/dataset/20200605-M-2019153LC-The First Affiliated Hospital of Guangzhou Medical University - Liang Wenhua/芯片住院信息对应 .xlsx",
            usecols=[0])  # 芯片名称
        patient_id = pd.read_excel(
            "E:/研一/Community-Detection/dataset/20200605-M-2019153LC-The First Affiliated Hospital of Guangzhou Medical University - Liang Wenhua/芯片住院信息对应 .xlsx",
            usecols=[1])  # 患者ID

        id = 'none'
        for n in range(0, len(chip_excel)):
            if q == chip_excel['芯片名称'][n]:
                id = patient_id['患者ID'][n]
        print('患者id：', id)

        # '''csv的边转换成txt文件'''
        cell_edge = pd.read_excel(file_a + big_file_name + '/' + Name + '_边表.csv', usecols=[0])
        data_1 = cell_edge['细胞的边'].str.split(',', expand=True)
        data_2 = data_1[0].str.split('(', expand=True)
        data_3 = data_1[1].str.split(')', expand=True)
        l = data_2.shape[0]

        '''社区发现'''

        # 读取包含细胞社区分配的CSV文件
        rows = []
        with open(file_a + big_file_name + '/' + Name + '_DAPI_path_view_社区.csv', newline='') as csvfile:
            csv_reader = csv.reader(csvfile)
            for row in csv_reader:
                rows.append(row)

        # 将数据转换为 DataFrame
        community = pd.DataFrame(rows)
        community.fillna(-1, inplace=True)

        '''数据统计'''

        node_kind = pd.read_excel(file_a + big_file_name + '/' + Name + '_各类细胞坐标.csv',
                                  usecols=[2])  # node_kind为节点细胞类型
        node_p = pd.read_excel(file_a + big_file_name + '/' + Name + '_各类细胞坐标.csv', usecols=[1])  # node_p为细胞坐标
        node_position = node_p['细胞坐标'].str.split(',', expand=True)
        node_position_x = node_position[0].str.split('(',
                                                     expand=True)  # node_position_x为社区节点x坐标矩阵 node_position_x[1][i]
        node_position_y = node_position[1].str.split(')',
                                                     expand=True)  # node_position_y为社区节点y坐标矩阵 node_position_y[0][i]
        # 创建细胞到社区的映射
        cell_to_community = {}
        for index, row in community.iterrows():
            cells = [cell for cell in row if cell != -1]  # 获取同一行中所有有效细胞（属于同一个社区）
            # 把cells中前面有空格的空格去掉
            cells = [cell.strip() for cell in cells]
            community_index = index  # 假设社区的索引是行号
            for cell in cells:
                cell_to_community[cell] = community_index

        if panel == 1:
            CD4_CD4_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result1/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[0], encoding='gbk', na_values=[''])
            CD20_CD20_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result1/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[1], encoding='gbk', na_values=[''])
            CD38_CD38_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result1/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[2], encoding='gbk', na_values=[''])
            CD66B_CD66B_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result1/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[3], encoding='gbk', na_values=[''])
            FOXP3_FOXP3_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result1/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[4], encoding='gbk', na_values=[''])
            CD4_CD20_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result1/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[5], encoding='gbk', na_values=[''])
            CD4_CD38_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result1/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[6], encoding='gbk', na_values=[''])
            CD4_CD66B_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result1/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[7], encoding='gbk', na_values=[''])
            CD4_FOXP3_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result1/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[8], encoding='gbk', na_values=[''])
            CD20_CD38_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result1/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[9], encoding='gbk', na_values=[''])
            CD20_CD66B_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result1/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[10], encoding='gbk', na_values=[''])
            CD20_FOXP3_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result1/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[11], encoding='gbk', na_values=[''])
            CD38_CD66B_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result1/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[12], encoding='gbk', na_values=[''])
            CD38_FOXP3_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result1/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[13], encoding='gbk', na_values=[''])
            CD66B_FOXP3_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result1/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[14], encoding='gbk', na_values=[''])

            if pd.isna(CD4_CD4_data['CD4-CD4连接边']).all():
                CD4_CD4_edge = []
            elif str(CD4_CD4_data['CD4-CD4连接边']) != 'nan' and str(CD4_CD4_data['CD4-CD4连接边']) != '':
                CD4_CD4_edge = CD4_CD4_data['CD4-CD4连接边'].str.split(';', expand=True)
                # 如果CD4_CD4_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD4_CD4_edge[0]
                col2 = CD4_CD4_edge[1]
                # 判断CD4_CD4_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD4_CD4_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD4_CD4_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD4_CD4_edge = []

            if pd.isna(CD20_CD20_data['CD20-CD20连接边']).all():
                CD20_CD20_edge = []
            elif str(CD20_CD20_data['CD20-CD20连接边']) != 'nan' and str(CD20_CD20_data['CD20-CD20连接边']) != '':
                CD20_CD20_edge = CD20_CD20_data['CD20-CD20连接边'].str.split(';', expand=True)
                # 如果CD20_CD20_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD20_CD20_edge[0]
                col2 = CD20_CD20_edge[1]
                # 判断CD20_CD20_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD20_CD20_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD20_CD20_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD20_CD20_edge = []

            if pd.isna(CD38_CD38_data['CD38-CD38连接边']).all():
                CD38_CD38_edge = []
            elif str(CD38_CD38_data['CD38-CD38连接边']) != 'nan' and str(CD38_CD38_data['CD38-CD38连接边']) != '':
                CD38_CD38_edge = CD38_CD38_data['CD38-CD38连接边'].str.split(';', expand=True)
                # 如果CD38_CD38_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD38_CD38_edge[0]
                col2 = CD38_CD38_edge[1]
                # 判断CD38_CD38_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD38_CD38_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD38_CD38_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD38_CD38_edge = []

            if pd.isna(CD66B_CD66B_data['CD66B-CD66B连接边']).all():
                CD66B_CD66B_edge = []
            elif str(CD66B_CD66B_data['CD66B-CD66B连接边']) != 'nan' and str(CD66B_CD66B_data['CD66B-CD66B连接边']) != '':
                CD66B_CD66B_edge = CD66B_CD66B_data['CD66B-CD66B连接边'].str.split(';', expand=True)
                # 如果CD66B_CD66B_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD66B_CD66B_edge[0]
                col2 = CD66B_CD66B_edge[1]
                # 判断CD66B_CD66B_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD66B_CD66B_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD66B_CD66B_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD66B_CD66B_edge = []

            if pd.isna(FOXP3_FOXP3_data['FOXP3-FOXP3连接边']).all():
                FOXP3_FOXP3_edge = []
            elif str(FOXP3_FOXP3_data['FOXP3-FOXP3连接边']) != 'nan' and str(FOXP3_FOXP3_data['FOXP3-FOXP3连接边']) != '':
                FOXP3_FOXP3_edge = FOXP3_FOXP3_data['FOXP3-FOXP3连接边'].str.split(';', expand=True)
                # 如果FOXP3_FOXP3_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = FOXP3_FOXP3_edge[0]
                col2 = FOXP3_FOXP3_edge[1]
                # 判断FOXP3_FOXP3_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            FOXP3_FOXP3_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        FOXP3_FOXP3_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                FOXP3_FOXP3_edge = []

            if pd.isna(CD4_CD20_data['CD4-CD20连接边']).all():
                CD4_CD20_edge = []
            elif str(CD4_CD20_data['CD4-CD20连接边']) != 'nan' and str(CD4_CD20_data['CD4-CD20连接边']) != '':
                CD4_CD20_edge = CD4_CD20_data['CD4-CD20连接边'].str.split(';', expand=True)
                # 如果CD4_CD20_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD4_CD20_edge[0]
                col2 = CD4_CD20_edge[1]
                # 判断CD4_CD20_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD4_CD20_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD4_CD20_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD4_CD20_edge = []

            if pd.isna(CD4_CD38_data['CD4-CD38连接边']).all():
                CD4_CD38_edge = []
            elif str(CD4_CD38_data['CD4-CD38连接边']) != 'nan' and str(CD4_CD38_data['CD4-CD38连接边']) != '':
                CD4_CD38_edge = CD4_CD38_data['CD4-CD38连接边'].str.split(';', expand=True)
                # 如果CD4_CD38_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD4_CD38_edge[0]
                col2 = CD4_CD38_edge[1]
                # 判断CD4_CD38_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD4_CD38_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD4_CD38_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD4_CD38_edge = []

            if pd.isna(CD4_CD66B_data['CD4-CD66B连接边']).all():
                CD4_CD66B_edge = []
            elif str(CD4_CD66B_data['CD4-CD66B连接边']) != 'nan' and str(CD4_CD66B_data['CD4-CD66B连接边']) != '':
                CD4_CD66B_edge = CD4_CD66B_data['CD4-CD66B连接边'].str.split(';', expand=True)
                # 如果CD4_CD66B_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD4_CD66B_edge[0]
                col2 = CD4_CD66B_edge[1]
                # 判断CD4_CD66B_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD4_CD66B_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD4_CD66B_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD4_CD66B_edge = []

            if pd.isna(CD4_FOXP3_data['CD4-FOXP3连接边']).all():
                CD4_FOXP3_edge = []
            elif str(CD4_FOXP3_data['CD4-FOXP3连接边']) != 'nan' and str(CD4_FOXP3_data['CD4-FOXP3连接边']) != '':
                CD4_FOXP3_edge = CD4_FOXP3_data['CD4-FOXP3连接边'].str.split(';', expand=True)
                # 如果CD4_FOXP3_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD4_FOXP3_edge[0]
                col2 = CD4_FOXP3_edge[1]
                # 判断CD4_FOXP3_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD4_FOXP3_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD4_FOXP3_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD4_FOXP3_edge = []


            if pd.isna(CD20_CD38_data['CD20-CD38连接边']).all():
                CD20_CD38_edge = []
            elif str(CD20_CD38_data['CD20-CD38连接边']) != 'nan' and str(CD20_CD38_data['CD20-CD38连接边']) != '':
                CD20_CD38_edge = CD20_CD38_data['CD20-CD38连接边'].str.split(';', expand=True)
                # 如果CD20_CD38_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD20_CD38_edge[0]
                col2 = CD20_CD38_edge[1]
                # 判断CD20_CD38_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD20_CD38_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD20_CD38_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD20_CD38_edge = []

            if pd.isna(CD20_CD66B_data['CD20-CD66B连接边']).all():
                CD20_CD66B_edge = []
            elif str(CD20_CD66B_data['CD20-CD66B连接边']) != 'nan' and str(CD20_CD66B_data['CD20-CD66B连接边']) != '':
                CD20_CD66B_edge = CD20_CD66B_data['CD20-CD66B连接边'].str.split(';', expand=True)
                # 如果CD20_CD66B_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD20_CD66B_edge[0]
                col2 = CD20_CD66B_edge[1]
                # 判断CD20_CD66B_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD20_CD66B_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD20_CD66B_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD20_CD66B_edge = []


            if pd.isna(CD20_FOXP3_data['CD20-FOXP3连接边']).all():
                CD20_FOXP3_edge = []
            elif str(CD20_FOXP3_data['CD20-FOXP3连接边']) != 'nan' and str(CD20_FOXP3_data['CD20-FOXP3连接边']) != '':
                CD20_FOXP3_edge = CD20_FOXP3_data['CD20-FOXP3连接边'].str.split(';', expand=True)
                # 如果CD20_FOXP3_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD20_FOXP3_edge[0]
                col2 = CD20_FOXP3_edge[1]
                # 判断CD20_FOXP3_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD20_FOXP3_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD20_FOXP3_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD20_FOXP3_edge = []

            if pd.isna(CD38_CD66B_data['CD38-CD66B连接边']).all():
                CD38_CD66B_edge = []
            elif str(CD38_CD66B_data['CD38-CD66B连接边']) != 'nan' and str(CD38_CD66B_data['CD38-CD66B连接边']) != '':
                CD38_CD66B_edge = CD38_CD66B_data['CD38-CD66B连接边'].str.split(';', expand=True)
                # 如果CD38_CD66B_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD38_CD66B_edge[0]
                col2 = CD38_CD66B_edge[1]
                # 判断CD38_CD66B_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD38_CD66B_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD38_CD66B_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD38_CD66B_edge = []

            if pd.isna(CD38_FOXP3_data['CD38-FOXP3连接边']).all():
                CD38_FOXP3_edge = []
            elif str(CD38_FOXP3_data['CD38-FOXP3连接边']) != 'nan' and str(CD38_FOXP3_data['CD38-FOXP3连接边']) != '':
                CD38_FOXP3_edge = CD38_FOXP3_data['CD38-FOXP3连接边'].str.split(';', expand=True)
                # 如果CD38_FOXP3_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD38_FOXP3_edge[0]
                col2 = CD38_FOXP3_edge[1]
                # 判断CD38_FOXP3_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD38_FOXP3_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD38_FOXP3_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD38_FOXP3_edge = []


            if pd.isna(CD66B_FOXP3_data['CD66B-FOXP3连接边']).all():
                CD66B_FOXP3_edge = []
            elif str(CD66B_FOXP3_data['CD66B-FOXP3连接边']) != 'nan' and str(
                    CD66B_FOXP3_data['CD66B-FOXP3连接边']) != '':
                CD66B_FOXP3_edge = CD66B_FOXP3_data['CD66B-FOXP3连接边'].str.split(';', expand=True)
                # 如果CD66B_FOXP3_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD66B_FOXP3_edge[0]
                col2 = CD66B_FOXP3_edge[1]
                # 判断CD66B_FOXP3_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD66B_FOXP3_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD66B_FOXP3_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD66B_FOXP3_edge = []

        else:
            CD8_CD8_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result2/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[0], encoding='gbk', na_values=[''])
            CD68_CD68_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result2/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[1], encoding='gbk', na_values=[''])
            CD133_CD133_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result2/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[2], encoding='gbk', na_values=[''])
            CD163_CD163_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result2/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[3], encoding='gbk', na_values=[''])
            PDL1_PDL1_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result2/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[4], encoding='gbk', na_values=[''])
            CD8_CD68_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result2/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[5], encoding='gbk', na_values=[''])
            CD8_CD133_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result2/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[6], encoding='gbk', na_values=[''])
            CD8_CD163_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result2/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[7], encoding='gbk', na_values=[''])
            CD8_PDL1_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result2/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[8], encoding='gbk', na_values=[''])
            CD68_CD133_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result2/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[9], encoding='gbk', na_values=[''])
            CD68_CD163_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result2/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[10], encoding='gbk', na_values=[''])
            CD68_PDL1_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result2/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[11], encoding='gbk', na_values=[''])
            CD133_CD163_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result2/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[12], encoding='gbk', na_values=[''])
            CD133_PDL1_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result2/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[13], encoding='gbk', na_values=[''])
            CD163_PDL1_data = pd.read_csv(
                'E:/研一/Community-Detection/dataset/result2/' + Name + '_DAPI_path_view.csv_各类细胞连接信息.csv',
                usecols=[14], encoding='gbk', na_values=[''])

            if pd.isna(CD8_CD8_data['CD8-CD8连接边']).all():
                CD8_CD8_edge = []
            elif str(CD8_CD8_data['CD8-CD8连接边']) != 'nan' and str(CD8_CD8_data['CD8-CD8连接边']) != '':
                CD8_CD8_edge = CD8_CD8_data['CD8-CD8连接边'].str.split(';', expand=True)
                # 如果CD8_CD8_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD8_CD8_edge[0]
                col2 = CD8_CD8_edge[1]
                # 判断CD8_CD8_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD8_CD8_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD8_CD8_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD8_CD8_edge = []

            if pd.isna(CD68_CD68_data['CD68-CD68连接边']).all():
                CD68_CD68_edge = []
            elif str(CD68_CD68_data['CD68-CD68连接边']) != 'nan' and str(CD68_CD68_data['CD68-CD68连接边']) != '':
                CD68_CD68_edge = CD68_CD68_data['CD68-CD68连接边'].str.split(';', expand=True)
                # 如果CD68_CD68_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD68_CD68_edge[0]
                col2 = CD68_CD68_edge[1]
                # 判断CD68_CD68_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD68_CD68_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD68_CD68_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD68_CD68_edge = []

            if pd.isna(CD133_CD133_data['CD133-CD133连接边']).all():
                CD133_CD133_edge = []
            elif str(CD133_CD133_data['CD133-CD133连接边']) != 'nan' and str(CD133_CD133_data['CD133-CD133连接边']) != '':
                CD133_CD133_edge = CD133_CD133_data['CD133-CD133连接边'].str.split(';', expand=True)
                # 如果CD133_CD133_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD133_CD133_edge[0]
                col2 = CD133_CD133_edge[1]
                # 判断CD133_CD133_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD133_CD133_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD133_CD133_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD133_CD133_edge = []

            if pd.isna(CD163_CD163_data['CD163-CD163连接边']).all():
                CD163_CD163_edge = []
            elif str(CD163_CD163_data['CD163-CD163连接边']) != 'nan' and str(CD163_CD163_data['CD163-CD163连接边']) != '':
                CD163_CD163_edge = CD163_CD163_data['CD163-CD163连接边'].str.split(';', expand=True)
                # 如果CD163_CD163_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD163_CD163_edge[0]
                col2 = CD163_CD163_edge[1]
                # 判断CD163_CD163_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD163_CD163_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD163_CD163_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD163_CD163_edge = []

            if pd.isna(PDL1_PDL1_data['PDL1-PDL1连接边']).all():
                PDL1_PDL1_edge = []
            elif str(PDL1_PDL1_data['PDL1-PDL1连接边']) != 'nan' and str(PDL1_PDL1_data['PDL1-PDL1连接边']) != '':
                PDL1_PDL1_edge = PDL1_PDL1_data['PDL1-PDL1连接边'].str.split(';', expand=True)
                # 如果PDL1_PDL1_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = PDL1_PDL1_edge[0]
                col2 = PDL1_PDL1_edge[1]
                # 判断PDL1_PDL1_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            PDL1_PDL1_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        PDL1_PDL1_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                PDL1_PDL1_edge = []

            if pd.isna(CD8_CD68_data['CD8-CD68连接边']).all():
                CD8_CD68_edge = []
            elif str(CD8_CD68_data['CD8-CD68连接边']) != 'nan' and str(CD8_CD68_data['CD8-CD68连接边']) != '':
                CD8_CD68_edge = CD8_CD68_data['CD8-CD68连接边'].str.split(';', expand=True)
                # 如果CD8_CD68_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD8_CD68_edge[0]
                col2 = CD8_CD68_edge[1]
                # 判断CD8_CD68_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD8_CD68_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD8_CD68_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD8_CD68_edge = []

            if pd.isna(CD8_CD133_data['CD8-CD133连接边']).all():
                CD8_CD133_edge = []
            elif str(CD8_CD133_data['CD8-CD133连接边']) != 'nan' and str(CD8_CD133_data['CD8-CD133连接边']) != '':
                CD8_CD133_edge = CD8_CD133_data['CD8-CD133连接边'].str.split(';', expand=True)
                # 如果CD8_CD133_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD8_CD133_edge[0]
                col2 = CD8_CD133_edge[1]
                # 判断CD8_CD133_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD8_CD133_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD8_CD133_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD8_CD133_edge = []

            if pd.isna(CD8_CD163_data['CD8-CD163连接边']).all():
                CD8_CD163_edge = []
            elif str(CD8_CD163_data['CD8-CD163连接边']) != 'nan' and str(CD8_CD163_data['CD8-CD163连接边']) != '':
                CD8_CD163_edge = CD8_CD163_data['CD8-CD163连接边'].str.split(';', expand=True)
                # 如果CD8_CD163_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD8_CD163_edge[0]
                col2 = CD8_CD163_edge[1]
                # 判断CD8_CD163_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD8_CD163_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD8_CD163_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD8_CD163_edge = []

            if pd.isna(CD8_PDL1_data['CD8-PDL1连接边']).all():
                CD8_PDL1_edge = []
            elif str(CD8_PDL1_data['CD8-PDL1连接边']) != 'nan' and str(CD8_PDL1_data['CD8-PDL1连接边']) != '':
                CD8_PDL1_edge = CD8_PDL1_data['CD8-PDL1连接边'].str.split(';', expand=True)
                # 如果CD8_PDL1_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD8_PDL1_edge[0]
                col2 = CD8_PDL1_edge[1]
                # 判断CD8_PDL1_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD8_PDL1_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD8_PDL1_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD8_PDL1_edge = []

            if pd.isna(CD68_CD133_data['CD68-CD133连接边']).all():
                CD68_CD133_edge = []
            elif str(CD68_CD133_data['CD68-CD133连接边']) != 'nan' and str(CD68_CD133_data['CD68-CD133连接边']) != '':
                CD68_CD133_edge = CD68_CD133_data['CD68-CD133连接边'].str.split(';', expand=True)
                # 如果CD68_CD133_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD68_CD133_edge[0]
                col2 = CD68_CD133_edge[1]
                # 判断CD68_CD133_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD68_CD133_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD68_CD133_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD68_CD133_edge = []

            if pd.isna(CD68_CD163_data['CD68-CD163连接边']).all():
                CD68_CD163_edge = []
            elif str(CD68_CD163_data['CD68-CD163连接边']) != 'nan' and str(CD68_CD163_data['CD68-CD163连接边']) != '':
                CD68_CD163_edge = CD68_CD163_data['CD68-CD163连接边'].str.split(';', expand=True)
                # 如果CD68_CD163_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD68_CD163_edge[0]
                col2 = CD68_CD163_edge[1]
                # 判断CD68_CD163_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD68_CD163_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD68_CD163_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD68_CD163_edge = []

            if pd.isna(CD68_PDL1_data['CD68-PDL1连接边']).all():
                CD68_PDL1_edge = []
            elif str(CD68_PDL1_data['CD68-PDL1连接边']) != 'nan' and str(CD68_PDL1_data['CD68-PDL1连接边']) != '':
                CD68_PDL1_edge = CD68_PDL1_data['CD68-PDL1连接边'].str.split(';', expand=True)
                # 如果CD68_PDL1_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD68_PDL1_edge[0]
                col2 = CD68_PDL1_edge[1]
                # 判断CD68_PDL1_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD68_PDL1_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD68_PDL1_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD68_PDL1_edge = []

            if pd.isna(CD133_CD163_data['CD133-CD163连接边']).all():
                CD133_CD163_edge = []
            elif str(CD133_CD163_data['CD133-CD163连接边']) != 'nan' and str(CD133_CD163_data['CD133-CD163连接边']) != '':
                CD133_CD163_edge = CD133_CD163_data['CD133-CD163连接边'].str.split(';', expand=True)
                # 如果CD133_CD163_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD133_CD163_edge[0]
                col2 = CD133_CD163_edge[1]
                # 判断CD133_CD163_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD133_CD163_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD133_CD163_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD133_CD163_edge = []

            if pd.isna(CD133_PDL1_data['CD133-PDL1连接边']).all():
                CD133_PDL1_edge = []
            elif str(CD133_PDL1_data['CD133-PDL1连接边']) != 'nan' and str(CD133_PDL1_data['CD133-PDL1连接边']) != '':
                CD133_PDL1_edge = CD133_PDL1_data['CD133-PDL1连接边'].str.split(';', expand=True)
                # 如果CD133_PDL1_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD133_PDL1_edge[0]
                col2 = CD133_PDL1_edge[1]
                # 判断CD133_PDL1_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD133_PDL1_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD133_PDL1_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD133_PDL1_edge = []

            if pd.isna(CD163_PDL1_data['CD163-PDL1连接边']).all():
                CD163_PDL1_edge = []
            elif str(CD163_PDL1_data['CD163-PDL1连接边']) != 'nan' and str(CD163_PDL1_data['CD163-PDL1连接边']) != '':
                CD163_PDL1_edge = CD163_PDL1_data['CD163-PDL1连接边'].str.split(';', expand=True)
                # 如果CD163_PDL1_edge中的两个点属于同一个社区，则保留。否则，删除该边。
                col1 = CD163_PDL1_edge[0]
                col2 = CD163_PDL1_edge[1]
                # 判断CD163_PDL1_edge同一行的两个点是否属于同一个社区,属于同一个社区，则保留。否则，删除该边。
                # 只取非nan的部分
                col1 = col1[~pd.isna(col1)].reset_index(drop=True)
                col2 = col2[~pd.isna(col2)].reset_index(drop=True)
                for i in range(len(col1)):
                    if col1[i] in cell_to_community and col2[i] in cell_to_community:
                        if cell_to_community[col1[i]] != cell_to_community[col2[i]]:
                            CD163_PDL1_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
                    else:
                        CD163_PDL1_edge.iloc[i, :] = np.nan  # 使用 np.nan 来代替空值
            else:
                CD163_PDL1_edge = []

            pass

        # 如果两个细胞属于同一个社区，则保留他们之间的连接边。如果两个细胞不属于同一个社区，则要删除他们之间的连接边。

        A_nodelist = []
        B_nodelist = []
        C_nodelist = []
        D_nodelist = []
        E_nodelist = []

        for i in range(community.shape[0]):  # 行数
            for j in range(community.shape[1]):  # 列数
                if int(community[j][i]) == -1:
                    continue  # 如果当前元素为 -1，则跳过，继续下一个元素的处理

                # 对非 -1 的元素执行其他操作
                node_type = node_kind['细胞类型'][int(community[j][i])]
                # if A == node_type:
                #     A_nodelist.append(int(community[j][i]))
                # elif B == node_type:
                #     B_nodelist.append(int(community[j][i]))
                # elif C == node_type:
                #     C_nodelist.append(int(community[j][i]))
                # elif D == node_type:
                #     D_nodelist.append(int(community[j][i]))
                # elif E == node_type:
                #     E_nodelist.append(int(community[j][i]))

                if 'CD4' == node_type:
                    A_nodelist.append(int(community[j][i]))
                elif 'CD20' == node_type:
                    B_nodelist.append(int(community[j][i]))
                elif 'CD38' == node_type:
                    C_nodelist.append(int(community[j][i]))
                elif 'CD66B' == node_type:
                    D_nodelist.append(int(community[j][i]))
                elif 'FOXP3' == node_type:
                    E_nodelist.append(int(community[j][i]))

        ws.cell(line, 1).value = Name
        ws.cell(line, 2).value = i + 1
        ws.cell(line, 43).value = id

        all_amount = len(A_nodelist) + len(B_nodelist) + len(C_nodelist) + len(D_nodelist) + len(E_nodelist)
        print("all_amount:", all_amount)
        '''计算各种类细胞的数目'''
        A_amount = len(A_nodelist)
        B_amount = len(B_nodelist)
        C_amount = len(C_nodelist)
        D_amount = len(D_nodelist)
        E_amount = len(E_nodelist)

        '''计算各种类细胞的百分比'''
        A_percent = (A_amount / all_amount) * 100
        B_percent = (B_amount / all_amount) * 100
        C_percent = (C_amount / all_amount) * 100
        D_percent = (D_amount / all_amount) * 100
        E_percent = (E_amount / all_amount) * 100

        '''写入各种类细胞的数目和百分比'''
        ws.cell(line, 3).value = A_amount / (i + 1)
        ws.cell(line, 4).value = B_amount / (i + 1)
        ws.cell(line, 5).value = C_amount / (i + 1)
        ws.cell(line, 6).value = D_amount / (i + 1)
        ws.cell(line, 7).value = E_amount / (i + 1)

        ws.cell(line, 8).value = A_percent / (i + 1)
        ws.cell(line, 9).value = B_percent / (i + 1)
        ws.cell(line, 10).value = C_percent / (i + 1)
        ws.cell(line, 11).value = D_percent / (i + 1)
        ws.cell(line, 12).value = E_percent / (i + 1)

        '''计算各细胞间的连接数及距离'''
        CD4_CD4_number = CD20_CD20_number = CD38_CD38_number = CD66B_CD66B_number = FOXP3_FOXP3_number = CD4_CD20_number \
            = CD4_CD38_number = CD4_CD66B_number = CD4_FOXP3_number = CD20_CD38_number = CD20_CD66B_number = CD20_FOXP3_number \
            = CD38_CD66B_number = CD38_FOXP3_number = CD66B_FOXP3_number = CD8_CD8_number = CD68_CD68_number \
            = CD133_CD133_number = CD163_CD163_number = PDL1_PDL1_number = CD8_CD68_number = CD8_CD133_number = CD8_CD163_number \
            = CD8_PDL1_number = CD68_CD133_number = CD68_CD163_number = CD68_PDL1_number = CD133_CD163_number = CD133_PDL1_number \
            = CD163_PDL1_number = 0

        CD4_CD4_dis = CD20_CD20_dis = CD38_CD38_dis = CD66B_CD66B_dis = FOXP3_FOXP3_dis = CD4_CD20_dis \
            = CD4_CD38_dis = CD4_CD66B_dis = CD4_FOXP3_dis = CD20_CD38_dis = CD20_CD66B_dis = CD20_FOXP3_dis \
            = CD38_CD66B_dis = CD38_FOXP3_dis = CD66B_FOXP3_dis = CD8_CD8_dis = CD68_CD68_dis \
            = CD133_CD133_dis = CD163_CD163_dis = PDL1_PDL1_dis = CD8_CD68_dis = CD8_CD133_dis = CD8_CD163_dis \
            = CD8_PDL1_dis = CD68_CD133_dis = CD68_CD163_dis = CD68_PDL1_dis = CD133_CD163_dis = CD133_PDL1_dis \
            = CD163_PDL1_dis = 0

        # panel1
        # for m in range(0, len(CD4_CD4_edge)):
        #     if str(CD4_CD4_edge[0][m]) != 'nan':
        #         CD4_CD4_number += 1
        #         k = int(CD4_CD4_edge[0][m])
        #         l = int(CD4_CD4_edge[1][m])
        #         dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
        #                 int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])))
        #         CD4_CD4_dis += dis  # ∣AB∣=√ [ (x1－x2)²+ (y1－y2)²]
        #
        # for m in range(0, len(CD20_CD20_edge)):
        #     if str(CD20_CD20_edge[0][m]) != 'nan':
        #         CD20_CD20_number += 1
        #         k = int(CD20_CD20_edge[0][m])
        #         l = int(CD20_CD20_edge[1][m])
        #         dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
        #                 int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])))
        #         CD20_CD20_dis += dis
        #
        # for m in range(0, len(CD38_CD38_edge)):
        #     if str(CD38_CD38_edge[0][m]) != 'nan':
        #         CD38_CD38_number += 1
        #         k = int(CD38_CD38_edge[0][m])
        #         l = int(CD38_CD38_edge[1][m])
        #         dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
        #                 int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])))
        #         CD38_CD38_dis += dis
        #
        # for m in range(0, len(CD66B_CD66B_edge)):
        #     if str(CD66B_CD66B_edge[0][m]) != 'nan':
        #         CD66B_CD66B_number += 1
        #         k = int(CD66B_CD66B_edge[0][m])
        #         l = int(CD66B_CD66B_edge[1][m])
        #         dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
        #                 int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])))
        #         CD66B_CD66B_dis += dis
        #
        # for m in range(0, len(FOXP3_FOXP3_edge)):
        #     if str(FOXP3_FOXP3_edge[0][m]) != 'nan':
        #         FOXP3_FOXP3_number += 1
        #         k = int(FOXP3_FOXP3_edge[0][m])
        #         l = int(FOXP3_FOXP3_edge[1][m])
        #         dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
        #                 int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])))
        #         FOXP3_FOXP3_dis += dis
        #
        # for m in range(0, len(CD4_CD20_edge)):
        #     if str(CD4_CD20_edge[0][m]) != 'nan':
        #         CD4_CD20_number += 1
        #         k = int(CD4_CD20_edge[0][m])
        #         l = int(CD4_CD20_edge[1][m])
        #         dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
        #                 int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])))
        #         CD4_CD20_dis += dis
        #
        # for m in range(0, len(CD4_CD38_edge)):
        #     if str(CD4_CD38_edge[0][m]) != 'nan':
        #         CD4_CD38_number += 1
        #         k = int(CD4_CD38_edge[0][m])
        #         l = int(CD4_CD38_edge[1][m])
        #         dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
        #                 int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])))
        #         CD4_CD38_dis += dis
        #
        # for m in range(0, len(CD4_CD66B_edge)):
        #     if str(CD4_CD66B_edge[0][m]) != 'nan':
        #         CD4_CD66B_number += 1
        #         k = int(CD4_CD66B_edge[0][m])
        #         l = int(CD4_CD66B_edge[1][m])
        #         dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
        #                 int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])))
        #         CD4_CD66B_dis += dis
        #
        # for m in range(0, len(CD4_FOXP3_edge)):
        #     if str(CD4_FOXP3_edge[0][m]) != 'nan':
        #         CD4_FOXP3_number += 1
        #         k = int(CD4_FOXP3_edge[0][m])
        #         l = int(CD4_FOXP3_edge[1][m])
        #         dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
        #                 int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])))
        #         CD4_FOXP3_dis += dis
        #
        # for m in range(0, len(CD20_CD38_edge)):
        #     if str(CD20_CD38_edge[0][m]) != 'nan':
        #         CD20_CD38_number += 1
        #         k = int(CD20_CD38_edge[0][m])
        #         l = int(CD20_CD38_edge[1][m])
        #         dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
        #                 int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])))
        #         CD20_CD38_dis += dis
        #
        # for m in range(0, len(CD20_CD66B_edge)):
        #     if str(CD20_CD66B_edge[0][m]) != 'nan':
        #         CD20_CD66B_number += 1
        #         k = int(CD20_CD66B_edge[0][m])
        #         l = int(CD20_CD66B_edge[1][m])
        #         dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
        #                 int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])))
        #         CD20_CD66B_dis += dis
        #
        # for m in range(0, len(CD20_FOXP3_edge)):
        #     if str(CD20_FOXP3_edge[0][m]) != 'nan':
        #         CD20_FOXP3_number += 1
        #         k = int(CD20_FOXP3_edge[0][m])
        #         l = int(CD20_FOXP3_edge[1][m])
        #         dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
        #                 int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])))
        #         CD20_FOXP3_dis += dis
        #
        # for m in range(0, len(CD38_CD66B_edge)):
        #     if str(CD38_CD66B_edge[0][m]) != 'nan':
        #         CD38_CD66B_number += 1
        #         k = int(CD38_CD66B_edge[0][m])
        #         l = int(CD38_CD66B_edge[1][m])
        #         dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
        #                 int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])))
        #         CD38_CD66B_dis += dis
        #
        # for m in range(0, len(CD38_FOXP3_edge)):
        #     if str(CD38_FOXP3_edge[0][m]) != 'nan':
        #         CD38_FOXP3_number += 1
        #         k = int(CD38_FOXP3_edge[0][m])
        #         l = int(CD38_FOXP3_edge[1][m])
        #         dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
        #                 int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])))
        #         CD38_FOXP3_dis += dis
        #
        # for m in range(0, len(CD66B_FOXP3_edge)):
        #     if str(CD66B_FOXP3_edge[0][m]) != 'nan':
        #         CD66B_FOXP3_number += 1
        #         k = int(CD66B_FOXP3_edge[0][m])
        #         l = int(CD66B_FOXP3_edge[1][m])
        #         dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
        #                 int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
        #                                 int(node_position_y[0][k]) - int(node_position_y[0][l])))
        #         CD66B_FOXP3_dis += dis

        # panel2
        for m in range(0, len(CD8_CD8_edge)):
            if str(CD8_CD8_edge[0][m]) != 'nan':
                CD8_CD8_number += 1
                k = int(CD8_CD8_edge[0][m])
                l = int(CD8_CD8_edge[1][m])
                dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
                        int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])))
                CD8_CD8_dis += dis  # ∣AB∣=√ [ (x1－x2)²+ (y1－y2)²]

        for m in range(0, len(CD68_CD68_edge)):
            if str(CD68_CD68_edge[0][m]) != 'nan':
                CD68_CD68_number += 1
                k = int(CD68_CD68_edge[0][m])
                l = int(CD68_CD68_edge[1][m])
                dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
                        int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])))
                CD68_CD68_dis += dis

        for m in range(0, len(CD133_CD133_edge)):
            if str(CD133_CD133_edge[0][m]) != 'nan':
                CD133_CD133_number += 1
                k = int(CD133_CD133_edge[0][m])
                l = int(CD133_CD133_edge[1][m])
                dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
                        int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])))
                CD133_CD133_dis += dis

        for m in range(0, len(CD163_CD163_edge)):
            if str(CD163_CD163_edge[0][m]) != 'nan':
                CD163_CD163_number += 1
                k = int(CD163_CD163_edge[0][m])
                l = int(CD163_CD163_edge[1][m])
                dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
                        int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])))
                CD163_CD163_dis += dis

        for m in range(0, len(PDL1_PDL1_edge)):
            if str(PDL1_PDL1_edge[0][m]) != 'nan':
                PDL1_PDL1_number += 1
                k = int(PDL1_PDL1_edge[0][m])
                l = int(PDL1_PDL1_edge[1][m])
                dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
                        int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])))
                PDL1_PDL1_dis += dis

        for m in range(0, len(CD8_CD68_edge)):
            if str(CD8_CD68_edge[0][m]) != 'nan':
                CD8_CD68_number += 1
                k = int(CD8_CD68_edge[0][m])
                l = int(CD8_CD68_edge[1][m])
                dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
                        int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])))
                CD8_CD68_dis += dis

        for m in range(0, len(CD8_CD133_edge)):
            if str(CD8_CD133_edge[0][m]) != 'nan':
                CD8_CD133_number += 1
                k = int(CD8_CD133_edge[0][m])
                l = int(CD8_CD133_edge[1][m])
                dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
                        int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])))
                CD8_CD133_dis += dis

        for m in range(0, len(CD8_CD163_edge)):
            if str(CD8_CD163_edge[0][m]) != 'nan':
                CD8_CD163_number += 1
                k = int(CD8_CD163_edge[0][m])
                l = int(CD8_CD163_edge[1][m])
                dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
                        int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])))
                CD8_CD163_dis += dis

        for m in range(0, len(CD8_PDL1_edge)):
            if str(CD8_PDL1_edge[0][m]) != 'nan':
                CD8_PDL1_number += 1
                k = int(CD8_PDL1_edge[0][m])
                l = int(CD8_PDL1_edge[1][m])
                dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
                        int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])))
                CD8_PDL1_dis += dis

        for m in range(0, len(CD68_CD133_edge)):
            if str(CD68_CD133_edge[0][m]) != 'nan':
                CD68_CD133_number += 1
                k = int(CD68_CD133_edge[0][m])
                l = int(CD68_CD133_edge[1][m])
                dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
                        int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])))
                CD68_CD133_dis += dis

        for m in range(0, len(CD68_CD163_edge)):
            if str(CD68_CD163_edge[0][m]) != 'nan':
                CD68_CD163_number += 1
                k = int(CD68_CD163_edge[0][m])
                l = int(CD68_CD163_edge[1][m])
                dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
                        int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])))
                CD68_CD163_dis += dis

        for m in range(0, len(CD68_PDL1_edge)):
            if str(CD68_PDL1_edge[0][m]) != 'nan':
                CD68_PDL1_number += 1
                k = int(CD68_PDL1_edge[0][m])
                l = int(CD68_PDL1_edge[1][m])
                dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
                        int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])))
                CD68_PDL1_dis += dis

        for m in range(0, len(CD133_CD163_edge)):
            if str(CD133_CD163_edge[0][m]) != 'nan':
                CD133_CD163_number += 1
                k = int(CD133_CD163_edge[0][m])
                l = int(CD133_CD163_edge[1][m])
                dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
                        int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])))
                CD133_CD163_dis += dis

        for m in range(0, len(CD133_PDL1_edge)):
            if str(CD133_PDL1_edge[0][m]) != 'nan':
                CD133_PDL1_number += 1
                k = int(CD133_PDL1_edge[0][m])
                l = int(CD133_PDL1_edge[1][m])
                dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
                        int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])))
                CD133_PDL1_dis += dis

        for m in range(0, len(CD163_PDL1_edge)):
            if str(CD163_PDL1_edge[0][m]) != 'nan':
                CD163_PDL1_number += 1
                k = int(CD163_PDL1_edge[0][m])
                l = int(CD163_PDL1_edge[1][m])
                dis = math.sqrt((int(node_position_x[1][k]) - int(node_position_x[1][l])) * (
                        int(node_position_x[1][k]) - int(node_position_x[1][l])) + (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])) * (
                                        int(node_position_y[0][k]) - int(node_position_y[0][l])))
                CD163_PDL1_dis += dis

        # panel1
        # ws.cell(line, 13).value = CD4_CD20_number / (i + 1)
        # if CD4_CD20_number != 0:
        #     ws.cell(line, 28).value = CD4_CD20_dis / CD4_CD20_number / (i + 1)
        # else:
        #     ws.cell(line, 28).value = 0
        #
        # ws.cell(line, 14).value = CD4_CD38_number / (i + 1)
        # if CD4_CD38_number != 0:
        #     ws.cell(line, 29).value = CD4_CD38_dis / CD4_CD38_number / (i + 1)
        # else:
        #     ws.cell(line, 29).value = 0
        #
        # ws.cell(line, 15).value = CD4_CD66B_number / (i + 1)
        # if CD4_CD66B_number != 0:
        #     ws.cell(line, 30).value = CD4_CD66B_dis / CD4_CD66B_number / (i + 1)
        # else:
        #     ws.cell(line, 30).value = 0
        #
        # ws.cell(line, 16).value = CD4_FOXP3_number / (i + 1)
        # if CD4_FOXP3_number != 0:
        #     ws.cell(line, 31).value = CD4_FOXP3_dis / CD4_FOXP3_number / (i + 1)
        # else:
        #     ws.cell(line, 31).value = 0
        #
        # ws.cell(line, 17).value = CD20_CD38_number / (i + 1)
        # if CD20_CD38_number != 0:
        #     ws.cell(line, 32).value = CD20_CD38_dis / CD20_CD38_number / (i + 1)
        # else:
        #     ws.cell(line, 32).value = 0
        #
        # ws.cell(line, 18).value = CD20_CD66B_number / (i + 1)
        # if CD20_CD66B_number != 0:
        #     ws.cell(line, 33).value = CD20_CD66B_dis / CD20_CD66B_number / (i + 1)
        # else:
        #     ws.cell(line, 33).value = 0
        #
        # ws.cell(line, 19).value = CD20_FOXP3_number / (i + 1)
        # if CD20_FOXP3_number != 0:
        #     ws.cell(line, 34).value = CD20_FOXP3_dis / CD20_FOXP3_number / (i + 1)
        # else:
        #     ws.cell(line, 34).value = 0
        #
        # ws.cell(line, 20).value = CD38_CD66B_number / (i + 1)
        # if CD38_CD66B_number != 0:
        #     ws.cell(line, 35).value = CD38_CD66B_dis / CD38_CD66B_number / (i + 1)
        # else:
        #     ws.cell(line, 35).value = 0
        #
        # ws.cell(line, 21).value = CD38_FOXP3_number / (i + 1)
        # if CD38_FOXP3_number != 0:
        #     ws.cell(line, 36).value = CD38_FOXP3_dis / CD38_FOXP3_number / (i + 1)
        # else:
        #     ws.cell(line, 36).value = 0
        #
        # ws.cell(line, 22).value = CD66B_FOXP3_number / (i + 1)
        # if CD66B_FOXP3_number != 0:
        #     ws.cell(line, 37).value = CD66B_FOXP3_dis / CD66B_FOXP3_number / (i + 1)
        # else:
        #     ws.cell(line, 37).value = 0
        #
        # ws.cell(line, 23).value = CD4_CD4_number / (i + 1)
        # if CD4_CD4_number != 0:
        #     ws.cell(line, 38).value = CD4_CD4_dis / CD4_CD4_number / (i + 1)
        # else:
        #     ws.cell(line, 38).value = 0
        #
        # ws.cell(line, 24).value = CD20_CD20_number / (i + 1)
        # if CD20_CD20_number != 0:
        #     ws.cell(line, 39).value = CD20_CD20_dis / CD20_CD20_number / (i + 1)
        # else:
        #     ws.cell(line, 39).value = 0
        #
        # ws.cell(line, 25).value = CD38_CD38_number / (i + 1)
        # if CD38_CD38_number != 0:
        #     ws.cell(line, 40).value = CD38_CD38_dis / CD38_CD38_number / (i + 1)
        # else:
        #     ws.cell(line, 40).value = 0
        #
        # ws.cell(line, 26).value = CD66B_CD66B_number / (i + 1)
        # if CD66B_CD66B_number != 0:
        #     ws.cell(line, 41).value = CD66B_CD66B_dis / CD66B_CD66B_number / (i + 1)
        # else:
        #     ws.cell(line, 41).value = 0
        #
        # ws.cell(line, 27).value = FOXP3_FOXP3_number / (i + 1)
        # if FOXP3_FOXP3_number != 0:
        #     ws.cell(line, 42).value = FOXP3_FOXP3_dis / FOXP3_FOXP3_number / (i + 1)
        # else:
        #     ws.cell(line, 42).value = 0

        # panel2
        ws.cell(line, 13).value = CD8_CD68_number / (i + 1)
        if CD8_CD68_number != 0:
            ws.cell(line, 28).value = CD8_CD68_dis / CD8_CD68_number / (i + 1)
        else:
            ws.cell(line, 28).value = 0

        ws.cell(line, 14).value = CD8_CD133_number / (i + 1)
        if CD8_CD133_number != 0:
            ws.cell(line, 29).value = CD8_CD133_dis / CD8_CD133_number / (i + 1)
        else:
            ws.cell(line, 29).value = 0

        ws.cell(line, 15).value = CD8_CD163_number / (i + 1)
        if CD8_CD163_number != 0:
            ws.cell(line, 30).value = CD8_CD163_dis / CD8_CD163_number / (i + 1)
        else:
            ws.cell(line, 30).value = 0

        ws.cell(line, 16).value = CD8_PDL1_number / (i + 1)
        if CD8_PDL1_number != 0:
            ws.cell(line, 31).value = CD8_PDL1_dis / CD8_PDL1_number / (i + 1)
        else:
            ws.cell(line, 31).value = 0

        ws.cell(line, 17).value = CD68_CD133_number / (i + 1)
        if CD68_CD133_number != 0:
            ws.cell(line, 32).value = CD68_CD133_dis / CD68_CD133_number / (i + 1)
        else:
            ws.cell(line, 32).value = 0

        ws.cell(line, 18).value = CD68_CD163_number / (i + 1)
        if CD68_CD163_number != 0:
            ws.cell(line, 33).value = CD68_CD163_dis / CD68_CD163_number / (i + 1)
        else:
            ws.cell(line, 33).value = 0

        ws.cell(line, 19).value = CD68_PDL1_number / (i + 1)
        if CD68_PDL1_number != 0:
            ws.cell(line, 34).value = CD68_PDL1_dis / CD68_PDL1_number / (i + 1)
        else:
            ws.cell(line, 34).value = 0

        ws.cell(line, 20).value = CD133_CD163_number / (i + 1)
        if CD133_CD163_number != 0:
            ws.cell(line, 35).value = CD133_CD163_dis / CD133_CD163_number / (i + 1)
        else:
            ws.cell(line, 35).value = 0

        ws.cell(line, 21).value = CD133_PDL1_number / (i + 1)
        if CD133_PDL1_number != 0:
            ws.cell(line, 36).value = CD133_PDL1_dis / CD133_PDL1_number / (i + 1)
        else:
            ws.cell(line, 36).value = 0

        ws.cell(line, 22).value = CD163_PDL1_number / (i + 1)
        if CD163_PDL1_number != 0:
            ws.cell(line, 37).value = CD163_PDL1_dis / CD163_PDL1_number / (i + 1)
        else:
            ws.cell(line, 37).value = 0

        ws.cell(line, 23).value = CD8_CD8_number / (i + 1)
        if CD8_CD8_number != 0:
            ws.cell(line, 38).value = CD8_CD8_dis / CD8_CD8_number / (i + 1)
        else:
            ws.cell(line, 38).value = 0

        ws.cell(line, 24).value = CD68_CD68_number / (i + 1)
        if CD68_CD68_number != 0:
            ws.cell(line, 39).value = CD68_CD68_dis / CD68_CD68_number / (i + 1)
        else:
            ws.cell(line, 39).value = 0

        ws.cell(line, 25).value = CD133_CD133_number / (i + 1)
        if CD133_CD133_number != 0:
            ws.cell(line, 40).value = CD133_CD133_dis / CD133_CD133_number / (i + 1)
        else:
            ws.cell(line, 40).value = 0

        ws.cell(line, 26).value = CD163_CD163_number / (i + 1)
        if CD163_CD163_number != 0:
            ws.cell(line, 41).value = CD163_CD163_dis / CD163_CD163_number / (i + 1)
        else:
            ws.cell(line, 41).value = 0

        ws.cell(line, 27).value = PDL1_PDL1_number / (i + 1)
        if PDL1_PDL1_number != 0:
            ws.cell(line, 42).value = PDL1_PDL1_dis / PDL1_PDL1_number / (i + 1)
        else:
            ws.cell(line, 42).value = 0

        line += 1

        # wb.save('E:/研一/Community-Detection/dataset/结构优化/Panel1/2.5/Panel1_社区优化特征_table1_cd.xlsx')
        # wb.save('E:/研一/Community-Detection/dataset/结构优化/Panel1/2.5/Panel1_社区优化特征_table2_cd.xlsx')

        # wb.save('E:/研一/Community-Detection/dataset/结构优化/Panel2/2.5/Panel2_社区优化特征_table1_cd.xlsx')
        # wb.save('E:/研一/Community-Detection/dataset/结构优化/Panel2/2.5/Panel2_社区优化特征_table2_cd.xlsx')
        wb.save('E:/研一/Community-Detection/dataset/结构优化/Panel2/2.5/Panel2_社区优化特征_table3_cd.xlsx')
