import openpyxl as ox
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.tri as tri
import cv2
import numpy as np
import pandas as pd
import math
import matplotlib.pyplot as plt
import re
import os


'''读取文件名函数'''
def file_name(file_dir):
    for root, dirs, files in os.walk(file_dir):
        return files

'''获取文件夹名字的函数'''
def getfilename(filename):
    for root, dirs, files in os.walk(filename):
        array = dirs
        if array:
            return array

'''找特定单元格的行数'''
def find_row(num_value, file_name):
    """
    Returns the row number based on the value of the specified cell
    """
    demo_df = pd.read_excel(file_name)
    for indexs in demo_df.index:
            if (str(demo_df.loc[indexs].values[0]) == num_value):
                row = str(indexs + 2).rstrip('L')
                return row

'''excel表格处理'''

file_a= 'D:/1-experiment/NewThreshold_data/20200605-M-2019153LC-The First Affiliated Hospital of Guangzhou Medical University - Liang Wenhua/Panel 2/117-3-T5-3/'

'''获取文件夹名字,生成列表并进行遍历'''
big_file =getfilename(file_a)
# print(big_file)

'''文件夹个数(文件夹列表中的序号)'''
big_file_num = 0


for big_file_name in big_file:
    print("循环第一个文件夹")
    print(big_file_name)


    '''去掉DATA字符，只留下前几位字符'''

    small_file = big_file_name[0:len(big_file_name) - 5]

    '''文件路径'''
    CD8_l=file_name(file_a+big_file_name+'/'+small_file+' FL/CD8/')
    CD68_l=file_name(file_a+big_file_name+'/'+small_file+' FL/CD68/')
    CD133_l=file_name(file_a+big_file_name+'/'+small_file+' FL/CD133/')
    CD163_l=file_name(file_a+big_file_name+'/'+small_file+' FL/CD163/')
    PDL1_l=file_name(file_a+big_file_name+'/'+small_file+' FL/PDL1/')
    csv_l=file_name(file_a+big_file_name+'/'+big_file_name+'/')

    '''文件名匹配并读取'''
    CD8_file_name = []
    CD68_file_name = []
    CD133_file_name = []
    CD163_file_name = []
    PDL1_file_name = []
    csv_file_name = []


    CD8_pattern = re.compile('C1-.+jpg')
    CD68_pattern = re.compile('C1-.+jpg')
    CD133_pattern = re.compile('C2-.+jpg')
    CD163_pattern = re.compile('C2-.+jpg')
    PDL1_pattern = re.compile('C1-.+jpg')
    csv_pattern = re.compile('.+csv')


    for CD8_i in CD8_l:
        CD8_result=CD8_pattern.findall(CD8_i)
        if(CD8_result!=[]):
            CD8_file_name.append(CD8_result)

    for CD68_i in CD68_l:
        CD68_result=CD68_pattern.findall(CD68_i)
        if(CD68_result!=[]):
            CD68_file_name.append(CD68_result)

    for CD133_i in CD133_l:
        CD133_result=CD133_pattern.findall(CD133_i)
        if(CD133_result!=[]):
            CD133_file_name.append(CD133_result)


    for CD163_i in CD163_l:
        CD163_result=CD163_pattern.findall(CD163_i)
        if(CD163_result!=[]):
            CD163_file_name.append(CD163_result)

    for PDL1_i in PDL1_l:
        PDL1_result=PDL1_pattern.findall(PDL1_i)
        if(PDL1_result!=[]):
            PDL1_file_name.append(PDL1_result)

    for csv_i in csv_l:
        csv_result=csv_pattern.findall(csv_i)
        if(csv_result!=[]):
            csv_file_name.append(csv_result)



    for CD8_jpg,CD68_jpg,CD133_jpg,CD163_jpg,PDL1_jpg,csv_name in zip(CD8_file_name,CD68_file_name,CD133_file_name,CD163_file_name,PDL1_file_name,csv_file_name):

        '''创建csv表'''
        wb = ox.Workbook()
        ws = wb.active

        ws['A1'] = '编号'
        ws['B1'] = '细胞坐标'
        ws['C1'] = '细胞类型'

        line_csv = 2

        '''csv文件名'''
        df_csv = pd.read_csv(file_a+big_file_name+'/'+big_file_name+'/'+str(csv_name[0]),usecols=[2])
        # print(csv_name[0])





        '''CD8坐标提取'''
        img_CD8 = cv2.imread(file_a+big_file_name+'/'+small_file+' FL/CD8/'+ str(CD8_jpg[0]))
        hsv_CD8 = cv2.cvtColor(img_CD8, cv2.COLOR_BGR2HSV)  # 色彩空间转换为hsv，分离.

        low_CD8 = np.array([0, 43, 46])
        high_CD8 = np.array([10, 255, 255])
        dst_CD8 = cv2.inRange(src=hsv_CD8, lowerb=low_CD8, upperb=high_CD8)  # HSV高低阈值，提取图像部分区域
        # 寻找白色的像素点坐标。
        # 白色像素值是255，所以np.where(dst==255)
        xy_CD8 = np.column_stack(np.where(dst_CD8 == 255))
        xy_CD8=xy_CD8[:, [1,0]]
        # print(xy)
        # df_CD8 = pd.read_csv("C:/Users/Administrator/Desktop/test/328573/S6-2 DATA/S6-2_Core[1,1,C]_[15214,40593]_DAPI_path_view.csv",usecols=[1])
        data_df_CD8 = df_csv[df_csv.columns[0]].str.split('-', expand=True)
        # print(data_df)
        np_df_x_CD8 = np.array(data_df_CD8[1])
        np_df_y_CD8 = np.array(data_df_CD8[0])
        values_x_CD8 = [int(x) for x in np_df_x_CD8]
        values_y_CD8 = [int(x) for x in np_df_y_CD8]
        a_CD8 = np.dstack((values_x_CD8,values_y_CD8))
        a_CD8 = np.squeeze(a_CD8)
        # print(np_df_x)
        a_CD8 = set([tuple(t) for t in a_CD8])
        xy_CD8 = set([tuple(t) for t in xy_CD8])
        matched_CD8 = list(a_CD8.intersection(xy_CD8))

        A = len(matched_CD8)
        # print(matched_CD8)
        # print(matched_CD8.shape)


        '''CD68坐标提取'''
        img_CD68 = cv2.imread(file_a+big_file_name+'/'+small_file+' FL/CD68/'+ str(CD68_jpg[0]))
        hsv_CD68 = cv2.cvtColor(img_CD68, cv2.COLOR_BGR2HSV)  # 色彩空间转换为hsv，分离.

        low_CD68 = np.array([0, 43, 46])
        high_CD68 = np.array([10, 255, 255])
        dst_CD68 = cv2.inRange(src=hsv_CD68, lowerb=low_CD68, upperb=high_CD68)  # HSV高低阈值，提取图像部分区域
        # 寻找白色的像素点坐标。
        # 白色像素值是255，所以np.where(dst==255)
        xy_CD68 = np.column_stack(np.where(dst_CD68 == 255))
        xy_CD68=xy_CD68[:, [1,0]]
        # print(xy)
        # df_CD68 = pd.read_csv("C:/Users/Administrator/Desktop/BING 4-1 DATA/BING 4-1 DATA/"+str(csv_name[0]),usecols=[1])
        data_df_CD68 = df_csv[df_csv.columns[0]].str.split('-', expand=True)
        # print(data_df)
        np_df_x_CD68 = np.array(data_df_CD68[1])
        np_df_y_CD68 = np.array(data_df_CD68[0])
        values_x_CD68 = [int(x) for x in np_df_x_CD68]
        values_y_CD68 = [int(x) for x in np_df_y_CD68]
        a_CD68 = np.dstack((values_x_CD68,values_y_CD68))
        a_CD68 = np.squeeze(a_CD68)
        # print(a)
        # print(np_df_x)
        a_CD68 = set([tuple(t) for t in a_CD68])
        xy_CD68 = set([tuple(t) for t in xy_CD68])
        matched_CD68 = list(a_CD68.intersection(xy_CD68))
        B = len(matched_CD68)
        # print(matched_CD68)
        # print(matched_CD68.shape)


        '''CD133坐标提取'''
        img_CD133 = cv2.imread(file_a+big_file_name+'/'+small_file+' FL/CD133/'+ str(CD133_jpg[0]))
        hsv_CD133 = cv2.cvtColor(img_CD133, cv2.COLOR_BGR2HSV)  # 色彩空间转换为hsv，分离.

        low_CD133 = np.array([35, 43, 46])
        high_CD133 = np.array([77, 255, 255])
        dst_CD133 = cv2.inRange(src=hsv_CD133, lowerb=low_CD133, upperb=high_CD133)  # HSV高低阈值，提取图像部分区域
        # 寻找白色的像素点坐标。
        # 白色像素值是255，所以np.where(dst==255)
        xy_CD133 = np.column_stack(np.where(dst_CD133 == 255))
        xy_CD133=xy_CD133[:, [1,0]]
        # print(xy)
        # df_CD133 = pd.read_csv("C:/Users/Administrator/Desktop/test/328573/S6-2 DATA/S6-2_Core[1,1,C]_[15214,40593]_DAPI_path_view.csv",usecols=[1])
        data_df_CD133 = df_csv[df_csv.columns[0]].str.split('-', expand=True)
        # print(data_df)
        np_df_x_CD133 = np.array(data_df_CD133[1])
        np_df_y_CD133 = np.array(data_df_CD133[0])
        values_x_CD133 = [int(x) for x in np_df_x_CD133]
        values_y_CD133 = [int(x) for x in np_df_y_CD133]
        a_CD133 = np.dstack((values_x_CD133,values_y_CD133))
        a_CD133 = np.squeeze(a_CD133)
        # print(a_CD133)
        # print(np_df_x_CD133)
        a_CD133 = set([tuple(t) for t in a_CD133])
        xy_CD133 = set([tuple(t) for t in xy_CD133])
        matched_CD133 = list(a_CD133.intersection(xy_CD133))
        C = len(matched_CD133)
        # print(matched_CD133)
        # print(matched_CD133.shape)


        '''CD163坐标提取'''
        img_CD163 = cv2.imread(file_a+big_file_name+'/'+small_file+' FL/CD163/'+ str(CD163_jpg[0]))
        hsv_CD163 = cv2.cvtColor(img_CD163, cv2.COLOR_BGR2HSV)  # 色彩空间转换为hsv，分离.

        low_CD163 = np.array([35, 43, 46])
        high_CD163 = np.array([77, 255, 255])
        dst_CD163 = cv2.inRange(src=hsv_CD163, lowerb=low_CD163, upperb=high_CD163)  # HSV高低阈值，提取图像部分区域
        # 寻找白色的像素点坐标。
        # 白色像素值是255，所以np.where(dst==255)
        xy_CD163 = np.column_stack(np.where(dst_CD163 == 255))
        xy_CD163=xy_CD163[:, [1,0]]
        # print(xy)
        # df_CD163 = pd.read_csv("C:/Users/Administrator/Desktop/test/328573/S6-2 DATA/S6-2_Core[1,1,C]_[15214,40593]_DAPI_path_view.csv",usecols=[1])
        data_df_CD163 = df_csv[df_csv.columns[0]].str.split('-', expand=True)
        # print(data_df)
        np_df_x_CD163 = np.array(data_df_CD163[1])
        np_df_y_CD163 = np.array(data_df_CD163[0])
        values_x_CD163 = [int(x) for x in np_df_x_CD163]
        values_y_CD163 = [int(x) for x in np_df_y_CD163]
        a_CD163 = np.dstack((values_x_CD163,values_y_CD163))
        a_CD163 = np.squeeze(a_CD163)
        # print(a_CD133)
        # print(np_df_x_CD133)
        a_CD163 = set([tuple(t) for t in a_CD163])
        xy_CD163 = set([tuple(t) for t in xy_CD163])
        matched_CD163 = list(a_CD163.intersection(xy_CD163))
        D = len(matched_CD163)
        # print(matched_CD163)
        # print(matched_CD163.shape)




        '''PDL1坐标提取'''
        img_PDL1 = cv2.imread(file_a+big_file_name+'/'+small_file+' FL/PDL1/'+ str(PDL1_jpg[0]))
        hsv_PDL1 = cv2.cvtColor(img_PDL1, cv2.COLOR_BGR2HSV)  # 色彩空间转换为hsv，分离.

        low_PDL1 = np.array([0, 43, 46])
        high_PDL1 = np.array([10, 255, 255])
        dst_PDL1 = cv2.inRange(src=hsv_PDL1, lowerb=low_PDL1, upperb=high_PDL1)  # HSV高低阈值，提取图像部分区域
        # 寻找白色的像素点坐标。
        # 白色像素值是255，所以np.where(dst==255)
        xy_PDL1 = np.column_stack(np.where(dst_PDL1 == 255))
        xy_PDL1=xy_PDL1[:, [1,0]]
        # print(xy)
        # df_PDL1 = pd.read_csv("C:/Users/Administrator/Desktop/test/328573/S6-2 DATA/S6-2_Core[1,1,C]_[15214,40593]_DAPI_path_view.csv",usecols=[1])
        data_df_PDL1 = df_csv [df_csv.columns[0]].str.split('-', expand=True)
        # print(data_df)
        np_df_x_PDL1 = np.array(data_df_PDL1[1])
        np_df_y_PDL1 = np.array(data_df_PDL1[0])
        values_x_PDL1 = [int(x) for x in np_df_x_PDL1]
        values_y_PDL1 = [int(x) for x in np_df_y_PDL1]
        a_PDL1 = np.dstack((values_x_PDL1,values_y_PDL1))
        a_PDL1 = np.squeeze(a_PDL1)
        # print(a_CD133)
        # print(np_df_x_CD133)
        a_PDL1= set([tuple(t) for t in a_PDL1])
        xy_PDL1 = set([tuple(t) for t in xy_PDL1])
        matched_PDL1 = list(a_PDL1.intersection(xy_PDL1))
        E = len(matched_PDL1)
        # print(matched_PDL1)
        # print(matched_PDL1.shape)
        # print(matched_CD8)


        '''各种类细胞合并连线'''
        cell_all_1 = []
        if(len(matched_CD8) != 0):
            cell_all_1 = matched_CD8

        if (len(matched_CD68) != 0):
            cell_all_2 = cell_all_1+matched_CD68


        if (len(matched_CD133) != 0):
            cell_all_3 = cell_all_2+matched_CD133


        if (len(matched_CD163) != 0):
            cell_all_4 = cell_all_3+matched_CD163


        if (len(matched_PDL1) != 0):
            cell_all_5 = cell_all_4 + matched_PDL1

        cell_all = np.array(list(cell_all_5))
        # print(matched_CD8.shape)
        # print(matched_CD133.shape)
        # print("细胞所有坐标：",cell_all)
        # print(cell_all.shape)
        # WIDTH = int(4028)
        # HEIGHT = int(3012)
        n = len(cell_all)  # n should be greater than 2
        #读取csv文件中的数据
        # df = pd.read_csv("C:/Users/Administrator/Desktop/10points.csv.",usecols=[0])
        # data_df = df[df.columns[0]].str.split('-', expand=True)
        # print(data_df)
        xs_all = cell_all[:,0]
        ys_all = cell_all[:,1]
        # print(xs_all)
        # print(ys_all)
        triang_all = tri.Triangulation(xs_all, ys_all)
        # print("总邻居表：",triang_all.edges)
        # print(len(triang_all.edges))




        '''各种类细胞进行连线'''
        # plt.figure(dpi=100, figsize=(WIDTH/100, HEIGHT/100))
        # plt.xlim(0, WIDTH)
        # plt.ylim(0, HEIGHT)
        # ax = plt.gca()#获取到当前坐标轴信息
        # ax.xaxis.set_ticks_position('top')   #将X坐标轴移到上面
        # ax.invert_yaxis()#反转Y坐标轴
        # ax.margins(0.1)
        # ax.set_aspect('equal')
        # ax.triplot(triang_all, 'b.-', linewidth=0.5)
        # ax.set_title('Plot of Delaunay triangulation')
        # plt.subplots_adjust(top=1,bottom=0,left=0,right=1,hspace=0,wspace=0)
        # plt.savefig('cell_all.jpg')



        '''合并所有细胞的邻居表'''
        #序号列表第一列
        list_a = []
        line1_list_x = []
        line1_list_y = []
        # print(triang_all.edges)
        for i in triang_all.edges:
            list_a.append(i[0])
        # print(list_a)
        for i in list_a:
            line1_list_x.append(xs_all[i])
        for i in list_a:
            line1_list_y.append(ys_all[i])
        # print(line1_list_x)
        # print(line1_list_y)
        line1_list = list(zip(line1_list_x,line1_list_y))
        # print(line1_list)

        #序号列表第二列
        list_b = []
        line2_list_x = []
        line2_list_y = []
        for i in triang_all.edges:
            list_b.append(i[1])
        # print(list_b)
        for i in list_b:
            line2_list_x.append(xs_all[i])
        for i in list_b:
            line2_list_y.append(ys_all[i])
        # print(line2_list_x)
        # print(line2_list_y)
        line2_list = list(zip(line2_list_x,line2_list_y))
        # print(line2_list)

        #合并坐标邻居列表
        line_list =  list(zip(line1_list,line2_list))
        # print("Line-list",line_list)
        #print(len(line_list))


        '''CD8：同种细胞之间的连接数'''
        # new_list_line1 = []   #将tuple元组转换为列表
        # new_list_line2 = []
        # print(len(line1_list))
        # print(line_list[0])
        # print(line_list[0][0])
        # # print(line_list[0][0][0])
        # for i in line_list:
        #     for a in i:
        #         print(a[0])
        #         print(a[1])
        CD8_list = []
        # print(list(line1_list[0]))

        for i in triang_all.edges:
            if list(line1_list[i[0]]) in matched_CD8:
                CD8_list.append(list(line1_list[i[0]]))
            if list(line2_list[i[0]]) in matched_CD8:
                CD8_list.append(list(line2_list[i[0]]))
        # print(cell_all)
        # print(cell_all)
        # print(cell_all.shape)
        #
        # print(matched_CD8)




        '''插入细胞坐标'''
        for i in matched_CD8:
            ws.cell(line_csv, 1).value = str(line_csv-2)
            ws.cell(line_csv, 2).value = str(i)
            ws.cell(line_csv, 3).value = str("CD8")
            line_csv += 1

        for i in matched_CD68:
            ws.cell(line_csv, 1).value = str(line_csv-2)
            ws.cell(line_csv, 2).value = str(i)
            ws.cell(line_csv, 3).value = str("CD68")
            line_csv += 1

        for i in matched_CD133:
            ws.cell(line_csv, 1).value = str(line_csv-2)
            ws.cell(line_csv, 2).value = str(i)
            ws.cell(line_csv, 3).value = str("CD133")
            line_csv += 1

        for i in matched_CD163:
            ws.cell(line_csv, 1).value = str(line_csv-2)
            ws.cell(line_csv, 2).value = str(i)
            ws.cell(line_csv, 3).value = str("CD163")
            line_csv += 1

        for i in matched_PDL1:
            ws.cell(line_csv, 1).value = str(line_csv-2)
            ws.cell(line_csv, 2).value = str(i)
            ws.cell(line_csv, 3).value = str("PDL1")
            line_csv += 1

        # '''插入excel表格平均距离'''
        # num_a = 0
        # for i in range(5, 48, 3):
        #     ws.cell(line_csv, i).value = list_ave[num_a]
        #     num_a += 1


        '''csv文件名拼接成芯片'''
        csv_pattern = re.compile('.+]')
        csv_result = str(csv_pattern.findall(csv_name[0])[0]) + "_各类细胞坐标.csv"



        wb.save(file_a+big_file_name+"/"+csv_result)

        '''创建边表的csv'''
        wb = ox.Workbook()
        ws = wb.active

        ws['A1'] = '细胞的边'
        ws['B1'] = '连成边的两个点坐标'

        line_csv_2 = 2
        '''边表存放'''
        for i,j in zip(triang_all.edges,line_list):
            # print(i)
            # print(j)
            ws.cell(line_csv_2, 1).value = "("+str(i[0])+","+str(i[1])+")"
            ws.cell(line_csv_2, 2).value = str(j[0])+","+str(j[1])
            line_csv_2 += 1
        csv_result_1 = str(csv_pattern.findall(csv_name[0])[0]) + "_边表.csv"
        wb.save(file_a + big_file_name + "/" + csv_result_1)








