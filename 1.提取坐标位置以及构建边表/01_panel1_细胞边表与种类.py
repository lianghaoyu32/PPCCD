import openpyxl as ox
import matplotlib.tri as tri
import cv2
import numpy as np
import pandas as pd
import math
import matplotlib.pyplot as plt
import re
import os


'''读取文件名函数'''
def file_name(file_dir):
    for root, dirs, files in os.walk(file_dir):
        return files

'''获取文件夹名字的函数'''
def getfilename(filename):
    for root, dirs, files in os.walk(filename):
        array = dirs
        if array:
            return array

'''找特定单元格的行数'''
def find_row(num_value, file_name):
    """
    Returns the row number based on the value of the specified cell
    """
    demo_df = pd.read_excel(file_name)
    for indexs in demo_df.index:
            if (str(demo_df.loc[indexs].values[0]) == num_value):
                row = str(indexs + 2).rstrip('L')
                return row

'''excel表格处理'''

file_a = 'E:/研一/Community-Detection/dataset/20200605-M-2019153LC-The First Affiliated Hospital of Guangzhou Medical University - Liang Wenhua/Panel 1/P2017004-P117-2-T5-1-20200217/'
# file_a= 'C:/Users/Administrator/Desktop/new/'

'''获取文件夹名字,生成列表并进行遍历'''
big_file =getfilename(file_a)
# print(big_file)

'''文件夹个数(文件夹列表中的序号)'''
big_file_num = 0


for big_file_name in big_file:
    print("循环第一个文件夹")
    print(big_file_name)


    '''去掉DATA字符，只留下前几位字符'''

    small_file = big_file_name[0:len(big_file_name) - 5]

    '''文件路径'''
    CD4_l=file_name(file_a+big_file_name+'/'+small_file+' FL/CD4/')
    CD20_l=file_name(file_a+big_file_name+'/'+small_file+' FL/CD20/')
    CD38_l=file_name(file_a+big_file_name+'/'+small_file+' FL/CD38/')
    CD66B_l=file_name(file_a+big_file_name+'/'+small_file+' FL/CD66B/')
    FOXP3_l=file_name(file_a+big_file_name+'/'+small_file+' FL/FOXP3/')
    csv_l=file_name(file_a+big_file_name+'/'+big_file_name+'/')

    '''文件名匹配并读取'''
    CD4_file_name = []
    CD20_file_name = []
    CD38_file_name = []
    CD66B_file_name = []
    FOXP3_file_name = []
    csv_file_name = []


    CD4_pattern = re.compile('C2-.+jpg')
    CD20_pattern = re.compile('C2-.+jpg')
    CD38_pattern = re.compile('C2-.+jpg')
    CD66B_pattern = re.compile('C2-.+jpg')
    FOXP3_pattern = re.compile('C1-.+jpg')
    csv_pattern = re.compile('.+csv')


    for CD4_i in CD4_l:
        CD4_result=CD4_pattern.findall(CD4_i)
        if(CD4_result!=[]):
            CD4_file_name.append(CD4_result)

    for CD20_i in CD20_l:
        CD20_result=CD20_pattern.findall(CD20_i)
        if(CD20_result!=[]):
            CD20_file_name.append(CD20_result)

    for CD38_i in CD38_l:
        CD38_result=CD38_pattern.findall(CD38_i)
        if(CD38_result!=[]):
            CD38_file_name.append(CD38_result)


    for CD66B_i in CD66B_l:
        CD66B_result=CD66B_pattern.findall(CD66B_i)
        if(CD66B_result!=[]):
            CD66B_file_name.append(CD66B_result)

    for FOXP3_i in FOXP3_l:
        FOXP3_result=FOXP3_pattern.findall(FOXP3_i)
        if(FOXP3_result!=[]):
            FOXP3_file_name.append(FOXP3_result)

    for csv_i in csv_l:
        csv_result=csv_pattern.findall(csv_i)
        if(csv_result!=[]):
            csv_file_name.append(csv_result)



    for CD4_jpg,CD20_jpg,CD38_jpg,CD66B_jpg,FOXP3_jpg,csv_name in zip(CD4_file_name,CD20_file_name,CD38_file_name,CD66B_file_name,FOXP3_file_name,csv_file_name):

        '''创建csv表'''
        wb = ox.Workbook()
        ws = wb.active

        ws['A1'] = '编号'
        ws['B1'] = '细胞坐标'
        ws['C1'] = '细胞类型'

        line_csv = 2

        '''csv文件名'''
        df_csv = pd.read_csv(file_a+big_file_name+'/'+big_file_name+'/'+str(csv_name[0]),usecols=[2])
        # print(csv_name[0])





        '''CD4坐标提取'''
        img_CD4 = cv2.imread(file_a+big_file_name+'/'+small_file+' FL/CD4/'+ str(CD4_jpg[0]))
        hsv_CD4 = cv2.cvtColor(img_CD4, cv2.COLOR_BGR2HSV)  # 色彩空间转换为hsv，分离.

        low_CD4 = np.array([35, 43, 46])
        high_CD4 = np.array([77, 255, 255])
        dst_CD4 = cv2.inRange(src=hsv_CD4, lowerb=low_CD4, upperb=high_CD4)  # HSV高低阈值，提取图像部分区域
        # 寻找白色的像素点坐标。
        # 白色像素值是255，所以np.where(dst==255)
        xy_CD4 = np.column_stack(np.where(dst_CD4 == 255))
        xy_CD4=xy_CD4[:, [1,0]]
        # print(xy)
        # df_CD4 = pd.read_csv("C:/Users/Administrator/Desktop/test/328573/S6-2 DATA/S6-2_Core[1,1,C]_[15214,40593]_DAPI_path_view.csv",usecols=[1])
        data_df_CD4 = df_csv[df_csv.columns[0]].str.split('-', expand=True)
        # data_df_CD4 = str(df_csv[df_csv.columns[0]]).split('-', expand=True)
        # print(data_df)
        np_df_x_CD4 = np.array(data_df_CD4[1])
        np_df_y_CD4 = np.array(data_df_CD4[0])
        values_x_CD4 = [int(x) for x in np_df_x_CD4]
        values_y_CD4 = [int(x) for x in np_df_y_CD4]
        a_CD4 = np.dstack((values_x_CD4,values_y_CD4))
        a_CD4 = np.squeeze(a_CD4)
        # print(a)
        # print(np_df_x)
        a_CD4 = set([tuple(t) for t in a_CD4])
        xy_CD4 = set([tuple(t) for t in xy_CD4])
        matched_CD4 = list(a_CD4.intersection(xy_CD4))

        A = len(matched_CD4)
        # print("CD4",matched_CD4)
        # print(matched_CD4.shape)


        '''CD20坐标提取'''
        img_CD20 = cv2.imread(file_a+big_file_name+'/'+small_file+' FL/CD20/'+ str(CD20_jpg[0]))
        hsv_CD20 = cv2.cvtColor(img_CD20, cv2.COLOR_BGR2HSV)  # 色彩空间转换为hsv，分离.

        low_CD20 = np.array([35, 43, 46])
        high_CD20 = np.array([77, 255, 255])
        dst_CD20 = cv2.inRange(src=hsv_CD20, lowerb=low_CD20, upperb=high_CD20)  # HSV高低阈值，提取图像部分区域
        # 寻找白色的像素点坐标。
        # 白色像素值是255，所以np.where(dst==255)
        xy_CD20 = np.column_stack(np.where(dst_CD20 == 255))
        xy_CD20=xy_CD20[:, [1,0]]
        # print(xy)
        # df_CD20 = pd.read_csv("C:/Users/Administrator/Desktop/BING 4-1 DATA/BING 4-1 DATA/"+str(csv_name[0]),usecols=[1])
        data_df_CD20 = df_csv[df_csv.columns[0]].str.split('-', expand=True)
        # print(data_df)
        np_df_x_CD20 = np.array(data_df_CD20[1])
        np_df_y_CD20 = np.array(data_df_CD20[0])
        values_x_CD20 = [int(x) for x in np_df_x_CD20]
        values_y_CD20 = [int(x) for x in np_df_y_CD20]
        a_CD20 = np.dstack((values_x_CD20,values_y_CD20))
        a_CD20 = np.squeeze(a_CD20)
        # print(a)
        # print(np_df_x)
        a_CD20 = set([tuple(t) for t in a_CD20])
        xy_CD20 = set([tuple(t) for t in xy_CD20])
        matched_CD20 = list(a_CD20.intersection(xy_CD20))
        B = len(matched_CD20)
        # print("CD20",matched_CD20)
        # print(matched_CD20.shape)


        '''CD38坐标提取'''
        img_CD38 = cv2.imread(file_a+big_file_name+'/'+small_file+' FL/CD38/'+ str(CD38_jpg[0]))
        hsv_CD38 = cv2.cvtColor(img_CD38, cv2.COLOR_BGR2HSV)  # 色彩空间转换为hsv，分离.

        low_CD38 = np.array([35, 43, 46])
        high_CD38 = np.array([77, 255, 255])
        dst_CD38 = cv2.inRange(src=hsv_CD38, lowerb=low_CD38, upperb=high_CD38)  # HSV高低阈值，提取图像部分区域
        # 寻找白色的像素点坐标。
        # 白色像素值是255，所以np.where(dst==255)
        xy_CD38 = np.column_stack(np.where(dst_CD38 == 255))
        xy_CD38=xy_CD38[:, [1,0]]
        # print(xy)
        # df_CD38 = pd.read_csv("C:/Users/Administrator/Desktop/test/328573/S6-2 DATA/S6-2_Core[1,1,C]_[15214,40593]_DAPI_path_view.csv",usecols=[1])
        data_df_CD38 = df_csv[df_csv.columns[0]].str.split('-', expand=True)
        # print(data_df)
        np_df_x_CD38 = np.array(data_df_CD38[1])
        np_df_y_CD38 = np.array(data_df_CD38[0])
        values_x_CD38 = [int(x) for x in np_df_x_CD38]
        values_y_CD38 = [int(x) for x in np_df_y_CD38]
        a_CD38 = np.dstack((values_x_CD38,values_y_CD38))
        a_CD38 = np.squeeze(a_CD38)
        # print(a_CD38)
        # print(np_df_x_CD38)
        a_CD38 = set([tuple(t) for t in a_CD38])
        xy_CD38 = set([tuple(t) for t in xy_CD38])
        matched_CD38 = list(a_CD38.intersection(xy_CD38))
        C = len(matched_CD38)
        # print("CD38",matched_CD38)
        # print(matched_CD38.shape)


        '''CD66B坐标提取'''
        img_CD66B = cv2.imread(file_a+big_file_name+'/'+small_file+' FL/CD66B/'+ str(CD66B_jpg[0]))
        hsv_CD66B = cv2.cvtColor(img_CD66B, cv2.COLOR_BGR2HSV)  # 色彩空间转换为hsv，分离.

        low_CD66B = np.array([35, 43, 46])
        high_CD66B = np.array([77, 255, 255])
        dst_CD66B = cv2.inRange(src=hsv_CD66B, lowerb=low_CD66B, upperb=high_CD66B)  # HSV高低阈值，提取图像部分区域
        # 寻找白色的像素点坐标。
        # 白色像素值是255，所以np.where(dst==255)
        xy_CD66B = np.column_stack(np.where(dst_CD66B == 255))
        xy_CD66B=xy_CD66B[:, [1,0]]
        # print(xy)
        # df_CD66B = pd.read_csv("C:/Users/Administrator/Desktop/test/328573/S6-2 DATA/S6-2_Core[1,1,C]_[15214,40593]_DAPI_path_view.csv",usecols=[1])
        data_df_CD66B = df_csv[df_csv.columns[0]].str.split('-', expand=True)
        # print(data_df)
        np_df_x_CD66B = np.array(data_df_CD66B[1])
        np_df_y_CD66B = np.array(data_df_CD66B[0])
        values_x_CD66B = [int(x) for x in np_df_x_CD66B]
        values_y_CD66B = [int(x) for x in np_df_y_CD66B]
        a_CD66B = np.dstack((values_x_CD66B,values_y_CD66B))
        a_CD66B = np.squeeze(a_CD66B)
        # print(a_CD38)
        # print(np_df_x_CD38)
        a_CD66B = set([tuple(t) for t in a_CD66B])
        xy_CD66B = set([tuple(t) for t in xy_CD66B])
        matched_CD66B = list(a_CD66B.intersection(xy_CD66B))
        D = len(matched_CD66B)
        # print("CD66B",matched_CD66B)
        # print(matched_CD66B.shape)




        '''FOXP3坐标提取'''
        img_FOXP3 = cv2.imread(file_a+big_file_name+'/'+small_file+' FL/FOXP3/'+ str(FOXP3_jpg[0]))
        hsv_FOXP3 = cv2.cvtColor(img_FOXP3, cv2.COLOR_BGR2HSV)  # 色彩空间转换为hsv，分离.

        low_FOXP3 = np.array([0, 43, 46])
        high_FOXP3 = np.array([10, 255, 255])
        dst_FOXP3 = cv2.inRange(src=hsv_FOXP3, lowerb=low_FOXP3, upperb=high_FOXP3)  # HSV高低阈值，提取图像部分区域
        # 寻找白色的像素点坐标。
        # 白色像素值是255，所以np.where(dst==255)
        xy_FOXP3 = np.column_stack(np.where(dst_FOXP3 == 255))
        xy_FOXP3=xy_FOXP3[:, [1,0]]
        # print(xy)
        # df_FOXP3 = pd.read_csv("C:/Users/Administrator/Desktop/test/328573/S6-2 DATA/S6-2_Core[1,1,C]_[15214,40593]_DAPI_path_view.csv",usecols=[1])
        data_df_FOXP3 = df_csv [df_csv.columns[0]].str.split('-', expand=True)
        # print(data_df)
        np_df_x_FOXP3 = np.array(data_df_FOXP3[1])
        np_df_y_FOXP3 = np.array(data_df_FOXP3[0])
        values_x_FOXP3 = [int(x) for x in np_df_x_FOXP3]
        values_y_FOXP3 = [int(x) for x in np_df_y_FOXP3]
        a_FOXP3 = np.dstack((values_x_FOXP3,values_y_FOXP3))
        a_FOXP3 = np.squeeze(a_FOXP3)
        # print(a_CD38)
        # print(np_df_x_CD38)
        a_FOXP3= set([tuple(t) for t in a_FOXP3])
        xy_FOXP3 = set([tuple(t) for t in xy_FOXP3])
        matched_FOXP3 = list(a_FOXP3.intersection(xy_FOXP3))
        E = len(matched_FOXP3)
        # print(matched_FOXP3)
        # print(matched_FOXP3.shape)
        # print(matched_CD4)


        '''各种类细胞合并连线'''
        cell_all_1 = []
        if(len(matched_CD4) != 0):
            cell_all_1 = matched_CD4

        if (len(matched_CD20) != 0):
            cell_all_2 = cell_all_1+matched_CD20
        else:
            cell_all_2 = cell_all_1

        if (len(matched_CD38) != 0):
            cell_all_3 = cell_all_2+matched_CD38
        else:
            cell_all_3 = cell_all_2

        if (len(matched_CD66B) != 0):
            cell_all_4 = cell_all_3+matched_CD66B
        else:
            cell_all_4 = cell_all_3

        if (len(matched_FOXP3) != 0):
            cell_all_5 = cell_all_4 + matched_FOXP3
        else:
            cell_all_5 = cell_all_4

        cell_all = np.array(list(cell_all_5))
        # print(matched_CD4.shape)
        # print(matched_CD38.shape)
        # print("细胞所有坐标：",cell_all)
        # print(cell_all.shape)
        WIDTH = int(4028)
        HEIGHT = int(3012)
        n = len(cell_all)  # n should be greater than 2
        #读取csv文件中的数据
        # df = pd.read_csv("C:/Users/Administrator/Desktop/10points.csv.",usecols=[0])
        # data_df = df[df.columns[0]].str.split('-', expand=True)
        # print(data_df)
        xs_all = cell_all[:,0]
        ys_all = cell_all[:,1]
        # print(xs_all)
        # print(ys_all)
        triang_all = tri.Triangulation(xs_all, ys_all)
        # print("总邻居表：",triang_all.edges)
        # print(len(triang_all.edges))




        '''各种类细胞进行连线'''
        # plt.figure(dpi=100, figsize=(WIDTH/100, HEIGHT/100))
        # plt.xlim(0, WIDTH)
        # plt.ylim(0, HEIGHT)
        # ax = plt.gca()#获取到当前坐标轴信息
        # ax.xaxis.set_ticks_position('top')   #将X坐标轴移到上面
        # ax.invert_yaxis()#反转Y坐标轴
        # ax.margins(0.1)
        # ax.set_aspect('equal')
        # ax.triplot(triang_all, 'b.-', linewidth=0.5)
        # ax.set_title('Plot of Delaunay triangulation')
        # plt.subplots_adjust(top=1,bottom=0,left=0,right=1,hspace=0,wspace=0)
        # plt.savefig('cell_all.jpg')



        '''合并所有细胞的邻居表'''
        #序号列表第一列
        list_a = []
        line1_list_x = []
        line1_list_y = []
        # print(triang_all.edges)
        for i in triang_all.edges:
            list_a.append(i[0])
        # print(list_a)
        for i in list_a:
            line1_list_x.append(xs_all[i])
        for i in list_a:
            line1_list_y.append(ys_all[i])
        # print(line1_list_x)
        # print(line1_list_y)
        line1_list = list(zip(line1_list_x,line1_list_y))
        # print(line1_list)

        #序号列表第二列
        list_b = []
        line2_list_x = []
        line2_list_y = []
        for i in triang_all.edges:
            list_b.append(i[1])
        # print(list_b)
        for i in list_b:
            line2_list_x.append(xs_all[i])
        for i in list_b:
            line2_list_y.append(ys_all[i])
        # print(line2_list_x)
        # print(line2_list_y)
        line2_list = list(zip(line2_list_x,line2_list_y))
        # print(line2_list)

        #合并坐标邻居列表
        line_list =  list(zip(line1_list,line2_list))
        # print("Line-list",line_list)
        #print(len(line_list))


        '''CD4：同种细胞之间的连接数'''
        # new_list_line1 = []   #将tuple元组转换为列表
        # new_list_line2 = []
        # print(len(line1_list))
        # print(line_list[0])
        # print(line_list[0][0])
        # # print(line_list[0][0][0])
        # for i in line_list:
        #     for a in i:
        #         print(a[0])
        #         print(a[1])
        CD4_list = []
        # print(list(line1_list[0]))

        for i in triang_all.edges:
            if list(line1_list[i[0]]) in matched_CD4:
                CD4_list.append(list(line1_list[i[0]]))
            if list(line2_list[i[0]]) in matched_CD4:
                CD4_list.append(list(line2_list[i[0]]))
        # print(cell_all)
        # print(cell_all)
        # print(cell_all.shape)
        #
        # print(matched_CD4)




        '''插入细胞坐标'''
        for i in matched_CD4:
            ws.cell(line_csv, 1).value = str(line_csv-2)
            ws.cell(line_csv, 2).value = str(i)
            ws.cell(line_csv, 3).value = str("CD4")
            line_csv += 1

        for i in matched_CD20:
            ws.cell(line_csv, 1).value = str(line_csv-2)
            ws.cell(line_csv, 2).value = str(i)
            ws.cell(line_csv, 3).value = str("CD20")
            line_csv += 1

        for i in matched_CD38:
            ws.cell(line_csv, 1).value = str(line_csv-2)
            ws.cell(line_csv, 2).value = str(i)
            ws.cell(line_csv, 3).value = str("CD38")
            line_csv += 1

        for i in matched_CD66B:
            ws.cell(line_csv, 1).value = str(line_csv-2)
            ws.cell(line_csv, 2).value = str(i)
            ws.cell(line_csv, 3).value = str("CD66B")
            line_csv += 1

        for i in matched_FOXP3:
            ws.cell(line_csv, 1).value = str(line_csv-2)
            ws.cell(line_csv, 2).value = str(i)
            ws.cell(line_csv, 3).value = str("FOXP3")
            line_csv += 1

        # '''插入excel表格平均距离'''
        # num_a = 0
        # for i in range(5, 48, 3):
        #     ws.cell(line_csv, i).value = list_ave[num_a]
        #     num_a += 1


        '''csv文件名拼接成芯片'''
        csv_pattern = re.compile('.+]')
        csv_result = str(csv_pattern.findall(csv_name[0])[0]) + "_各类细胞坐标.csv"



        wb.save(file_a+big_file_name+"/"+csv_result)

        '''创建边表的csv'''
        wb = ox.Workbook()
        ws = wb.active

        ws['A1'] = '细胞的边'
        ws['B1'] = '连成边的两个点坐标'

        line_csv_2 = 2
        '''边表存放'''
        for i,j in zip(triang_all.edges,line_list):
            # print(i)
            # print(j)
            ws.cell(line_csv_2, 1).value = "("+str(i[0])+","+str(i[1])+")"
            ws.cell(line_csv_2, 2).value = str(j[0])+","+str(j[1])
            line_csv_2 += 1
        csv_result_1 = str(csv_pattern.findall(csv_name[0])[0]) + "_边表.csv"
        wb.save(file_a + big_file_name + "/" + csv_result_1)








